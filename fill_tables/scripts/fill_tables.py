import os
import openpyxl
import pandas as pd
import gspread
import requests
import google.auth
import subprocess
import datetime
from datetime import date
from google.oauth2 import service_account
from gspread_pandas import Spread, Client 
from gspread_formatting import *

#Основная функция 
def main():
    print('Starting to fill tables')
    fill_table()
# Функция заполнения таблиц
def fill_table():
    try:
        table = Spread("PnL New",create_spread=False)
        print("Successfull authorized")
    except requests.exceptions.RequestException:
        print("Can't connect to Google API")
        raise SystemExit()
    reports = load_reports()
    for report in reports:
        print('Working with {}'.format(report.burse))
        fill_month(table,report)
        fill_Pertrader_month(table,report)
        fill_balances(table,report)
    check(table,report)
    print("Table filled")
# Функция загрузки отчётов из таблиц
def load_reports():
    reports = []
    reports_counter = 0
    reports_folder = os.listdir(os.getcwd() + '/reports')
    for report_file in reports_folder:
        if report_file.find('result.xlsx') != -1:
            try:
                workbook = openpyxl.load_workbook(filename = "./reports/" + report_file)
            except requests.exceptions.RequestException:
                print("Can't open report " + report_file)
                raise SystemExit()
            reports.append(Report(str(report_file),workbook))
            reports_counter += 1
    print("Reports loaded")
    return reports
# Заполнение листа месяца
def fill_month(table,report):

    #Определение диапазона копируемых ячеек
    report_sheet = report.workbook.get_sheet_by_name('Total')
    report_row = 2
    first_report_col = 1
    last_report_col = first_report_col + 3
    copied_row = report_sheet[number_to_fucking_a1(first_report_col) + str(report_row):number_to_fucking_a1(last_report_col) + str(report_row)]
    flat_copied_row = []
    for tup in copied_row:
        for t in tup:
            flat_copied_row.append(t.value)

    #Определение диапазона заполняемых ячеек
    sheet_month = table.find_sheet(get_month_from_numbers(report.month) + ' ' + str(report.year))
    filling_row = sheet_month.find(report.get_date()).row
    if (report.session == 'B'):
        first_filling_col = sheet_month.find(report.burse).col 
    else:
        first_filling_col = sheet_month.find(report.burse + ' Night').col 
    last_filling_col = first_filling_col + last_report_col - first_report_col
    try:
        table.update_cells((filling_row,first_filling_col),(filling_row,last_filling_col),list(flat_copied_row),sheet_month)
    except requests.exceptions.RequestException:
        print("Cant fill cell range in Month sheet")
    print('Month sheet of {} succesfully filled'.format(report.burse))
#Заполнение листа трейдеров
def fill_Pertrader_month(table,report):

    #Определение диапазона копируемых ячеек
    report_sheet = report.workbook.get_sheet_by_name('Per_trader')
    first_report_row = 2
    first_report_col = 1
    last_report_row = report_sheet.max_row
    last_report_col = first_report_col + 4
    copied_range = report_sheet[number_to_fucking_a1(first_report_col) + str(first_report_row):number_to_fucking_a1(last_report_col) + str(last_report_row)]
    flat_copied_range = []
    for tup in copied_range:
        for t in tup:
            flat_copied_range.append(t.value)
    #print("Copying cells from range {}{}:{}{}".format(number_to_fucking_a1(first_report_col),first_report_row,number_to_fucking_a1(last_report_col),last_report_row))

    #Определение диапазона заполняемых ячеек
    sheet_PerTrader = table.find_sheet('PerTrader_' + str(get_month_from_numbers(report.month)[0:3]))
    first_filling_col = sheet_PerTrader.find('session').col
    session_column = sheet_PerTrader.col_values(first_filling_col)
    first_filling_row = len(session_column) + 1
    last_filling_col = first_filling_col + last_report_col - first_report_col
    last_filling_row = first_filling_row + last_report_row - first_report_row
    filling_range = range_to_a1(first_filling_row,first_filling_col,last_filling_row,last_filling_col)
    #print('Copiing to range {}{}:{}{}'.format(number_to_fucking_a1(first_filling_col),first_filling_row,number_to_fucking_a1(last_filling_col - 1),last_filling_row - 1))
    
    #Заполнение таблицы
    try:
        table.update_cells((first_filling_row,first_filling_col),(last_filling_row,last_filling_col),list(flat_copied_range),sheet_PerTrader)
    except requests.exceptions.RequestException:
        print("Cant fill cell range in Per_Trader sheet")

    #Форматирование заполненного диапазона
    col_counter = first_filling_col
    while col_counter <= last_filling_col:
        fmt = get_effective_format(sheet_PerTrader,number_to_fucking_a1(col_counter) + str(first_filling_row-1))
        format_cell_range(sheet_PerTrader, '{}{}:{}{}'.format(number_to_fucking_a1(col_counter),str(first_filling_row),number_to_fucking_a1(col_counter),str(last_filling_row)),fmt)
        col_counter += 1
    print('PerTrader sheet of {} successfully filled'.format(report.burse))
#Заполнение листа балансов
def fill_balances(table,report):

    #Определение диапазона копируемых ячеек
    report_sheet = report.workbook.get_sheet_by_name('Balances')
    first_report_row = 2
    first_report_col = 1
    last_report_row = report_sheet.max_row
    last_report_col = first_report_col + 5
    copied_range = report_sheet[number_to_fucking_a1(first_report_col) + str(first_report_row):number_to_fucking_a1(last_report_col) + str(last_report_row)]
    flat_copied_range = []
    for tup in copied_range:
        for t in tup:
            flat_copied_range.append(t.value)

    #Определение диапазона заполняемых ячеек
    sheet_Balances = table.find_sheet('Balances')
    first_filling_col = 1
    session_column = sheet_Balances.col_values(first_filling_col)
    first_filling_row = len(session_column) + 1
    last_filling_col = first_filling_col + last_report_col - first_report_col
    last_filling_row = first_filling_row + last_report_row - first_report_row
    
    # Заполнение таблицы
    try:
        table.update_cells((first_filling_row,first_filling_col),(last_filling_row,last_filling_col),list(flat_copied_range),sheet_Balances)
    except requests.exceptions.RequestException:
        print("Cant fill cell range in Balances sheet")
    col_counter = first_filling_col
    while col_counter <= last_filling_col:
        fmt = get_effective_format(sheet_Balances,number_to_fucking_a1(col_counter) + str(first_filling_row - 1))
        if fmt.numberFormat:
            fmt.numberFormat.type = None
        #print('Formatting {}{}:{}{}'.format(number_to_fucking_a1(col_counter),str(first_filling_row),number_to_fucking_a1(col_counter),str(last_filling_row)))
        format_cell_range(sheet_Balances, '{}{}:{}{}'.format(number_to_fucking_a1(col_counter), str(first_filling_row), number_to_fucking_a1(col_counter), str(last_filling_row)), fmt)
        col_counter += 1
    print('Balances sheet of {} successfully filled'.format(report.burse))
#Вспомогательная функция разбиения имени отчёта, для получения инфы по нему
def parse_report_name(report_file):
    try:
        current_year = report_file[:4]
        current_month = report_file[4:6]
        current_day = report_file[6:8]
        burse = ''
        counter = 9
        while report_file[counter] != '_':
            burse += report_file[counter]
            counter += 1
        session = report_file[counter + 1:counter + 2]
    except requests.exceptions.RequestException:
        print("Can't parse filename")
        SystemExit()
    print('Open report: ' + 'Year ' + current_year + ' Month ' + current_month + ' day ' + current_day + ' burse ' + burse + ' session ' + session)
    return current_year, current_month, current_day, burse, session
#Класс отчёта
class Report():
    year = 0
    month = 0
    day = 0
    burse = ''
    session = ''
    workbook = openpyxl.Workbook()
    def __init__(self,report_name,workbook):
        self.year, self.month, self.day, self.burse, self.session = parse_report_name(report_name)
        self.workbook = workbook
    def get_date(self):
        return str(self.year + '.' + self.month + '.' + self.day)
#Получение строки месяца по числу
def get_month_from_numbers(number):
    numbers_to_string_month = {
        '01' : 'January',
        '02' : 'February',
        '03' : 'March',
        '04' : 'April',
        '05' : 'May',
        '06' : 'June',
        '07' : 'July',
        '08' : 'August',
        '09' : 'September',
        '10' : 'October',
        '11' : 'November',
        '12' : 'December'
    }
    return numbers_to_string_month[number]
#Поиск первой пустой строки в таблице
def get_empty_raw(session_col):
    str_list = filter(None,session_col)
    return len(str_list) + 1
# Приведение нотации таблицы к формату A1
def number_to_fucking_a1(number):
    number_to_a1 = {
        0 : '',
        1 : 'A',
        2 : 'B',
        3 : 'C',
        4 : 'D',
        5 : 'E',
        6 : 'F',
        7 : 'G',
        8 : 'H',
        9 : 'I',
        10 : 'J',
        11 : 'K',
        12 : 'L',
        13 : 'M',
        14 : 'N',
        15 : 'O',
        16 : 'P',
        17 : 'Q',
        18 : 'R',
        19 : 'S',
        20 : 'T',
        21 : 'U',
        22 : 'V',
        23 : 'W',
        24 : 'X',
        25 : 'Y',
        26 : 'Z'
    }
    a1_notation = ''
    while number // 26 > 0:
        a1_notation += number_to_a1[number // 26]
        number = number % 26
    a1_notation += number_to_a1[number]
    return a1_notation
# Преобразование диапазона ячеек к А1 нотации
def range_to_a1(first_row, first_col, last_row, last_col):
    #start = number_to_fucking_a1(first_col) + str(first_row)
    #end = number_to_fucking_a1(last_col) + str(last_row)
    range = number_to_fucking_a1(first_col) + str(first_row) + ':' + number_to_fucking_a1(last_col) + str(last_row)
    return range
# Функция проверки значения Check и Double Check
def check(table, report):
    summary_sheet = table.find_sheet('Summary_' + str(get_month_from_numbers(report.month)[0:3]))
    check_value = summary_sheet.cell(summary_sheet.find('Check').row,summary_sheet.find('Check').col + 1).value
    double_check_value = summary_sheet.cell(summary_sheet.find('Double Check').row,summary_sheet.find('Double Check').col + 1).value
    print('Check is {} and double check is {}'.format(check_value, double_check_value))
    if abs(float(check_value.replace(',','.'))) >= 10:
        print('WARNING!!! Check is ' + check_value)
    if abs(float(double_check_value.replace(',','.'))) >= 10:
        print('WARNING!!! Double Check is ' + double_check_value)

if __name__ == '__main__':
    main()